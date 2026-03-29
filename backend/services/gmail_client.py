import base64
import csv
from io import BytesIO, StringIO
import os
import re
from pathlib import Path
from typing import Any, Dict, List, Optional


class GmailClient:
    """Thin Gmail API wrapper for watch + message retrieval."""

    scope = 'https://www.googleapis.com/auth/gmail.readonly'

    def __init__(self, service: Any = None):
        self._service = service

    def start_watch(self, topic_name: str, label_ids: Optional[List[str]] = None) -> Dict[str, Any]:
        service = self._get_service()
        body = {
            'labelIds': label_ids or ['INBOX'],
            'topicName': topic_name,
        }
        return service.users().watch(userId='me', body=body).execute()

    def list_new_message_ids(self, start_history_id: str) -> List[str]:
        service = self._get_service()
        response = service.users().history().list(
            userId='me',
            startHistoryId=start_history_id,
            historyTypes=['messageAdded'],
        ).execute()

        message_ids: List[str] = []
        for history_item in response.get('history', []):
            for added in history_item.get('messagesAdded', []):
                message = added.get('message', {})
                msg_id = message.get('id')
                if msg_id and msg_id not in message_ids:
                    message_ids.append(msg_id)
        return message_ids

    def list_recent_message_ids(
        self,
        max_results: int = 1,
        label_ids: Optional[List[str]] = None,
        prefer_unread: bool = True,
    ) -> List[str]:
        service = self._get_service()
        effective_max = max(1, int(max_results))
        effective_labels = label_ids or ['INBOX']

        if prefer_unread:
            unread_response = service.users().messages().list(
                userId='me',
                maxResults=effective_max,
                labelIds=effective_labels,
                q='is:unread',
            ).execute()
            unread_message_ids = self._extract_message_ids_from_list_response(unread_response)
            if unread_message_ids:
                return unread_message_ids

        response = service.users().messages().list(
            userId='me',
            maxResults=effective_max,
            labelIds=effective_labels,
        ).execute()
        return self._extract_message_ids_from_list_response(response)

    def _extract_message_ids_from_list_response(self, response: Dict[str, Any]) -> List[str]:
        message_ids: List[str] = []
        for item in response.get('messages', []):
            msg_id = str(item.get('id', '')).strip()
            if msg_id and msg_id not in message_ids:
                message_ids.append(msg_id)
        return message_ids

    def get_email(self, msg_id: str) -> Dict[str, Any]:
        service = self._get_service()
        msg = service.users().messages().get(userId='me', id=msg_id, format='full').execute()

        payload = msg.get('payload', {})
        snippet = str(msg.get('snippet', '')).strip()
        headers = payload.get('headers', [])
        subject = self._header_value(headers, 'Subject')
        sender = self._header_value(headers, 'From')
        body = self._extract_body(payload)
        attachments = self._collect_attachments(payload=payload, service=service, message_id=msg_id)
        attachment_text = '\n\n'.join(
            chunk for chunk in (item.get('text_preview', '') for item in attachments) if chunk
        ).strip()

        if not body and snippet:
            body = snippet

        return {
            'message_id': msg.get('id', msg_id),
            'history_id': str(msg.get('historyId', '')),
            'subject': subject,
            'sender': sender,
            'body': body,
            'snippet': snippet,
            'attachments': attachments,
            'attachment_text': attachment_text,
        }

    def _get_service(self) -> Any:
        if self._service is not None:
            return self._service

        try:
            from google.auth.transport.requests import Request
            from google.oauth2.credentials import Credentials
            from google_auth_oauthlib.flow import InstalledAppFlow
            from googleapiclient.discovery import build
        except ImportError as exc:
            raise RuntimeError(
                'google-api-python-client, google-auth-httplib2, and google-auth-oauthlib are required for Gmail integration.'
            ) from exc

        token_path = Path(os.getenv('GMAIL_TOKEN_FILE', str(Path(__file__).resolve().parents[2] / 'data' / 'gmail_token.json')))
        credentials_path = os.getenv('GMAIL_CREDENTIALS_FILE', '').strip()
        client_id = os.getenv('GMAIL_CLIENT_ID', '').strip()
        client_secret = os.getenv('GMAIL_CLIENT_SECRET', '').strip()

        creds = None
        if token_path.exists():
            creds = Credentials.from_authorized_user_file(str(token_path), [self.scope])

        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())

        if not creds or not creds.valid:
            if credentials_path:
                flow = InstalledAppFlow.from_client_secrets_file(credentials_path, [self.scope])
            elif client_id and client_secret:
                flow = InstalledAppFlow.from_client_config(
                    self._oauth_client_config(client_id=client_id, client_secret=client_secret),
                    [self.scope],
                )
            else:
                raise RuntimeError(
                    'Set GMAIL_CREDENTIALS_FILE or GMAIL_CLIENT_ID/GMAIL_CLIENT_SECRET for Gmail OAuth.'
                )
            creds = flow.run_local_server(port=0)
            token_path.parent.mkdir(parents=True, exist_ok=True)
            token_path.write_text(creds.to_json(), encoding='utf-8')

        self._service = build('gmail', 'v1', credentials=creds)
        return self._service

    def _header_value(self, headers: List[Dict[str, Any]], name: str) -> str:
        for item in headers:
            if str(item.get('name', '')).lower() == name.lower():
                return str(item.get('value', ''))
        return ''

    def _oauth_client_config(self, client_id: str, client_secret: str) -> Dict[str, Any]:
        return {
            'installed': {
                'client_id': client_id,
                'client_secret': client_secret,
                'auth_uri': 'https://accounts.google.com/o/oauth2/auth',
                'token_uri': 'https://oauth2.googleapis.com/token',
                'redirect_uris': ['http://localhost'],
            }
        }

    def _extract_body(self, payload: Dict[str, Any]) -> str:
        plain = self._decode_part(payload, 'text/plain')
        if plain:
            return plain

        html = self._decode_part(payload, 'text/html')
        if html:
            return self._strip_html(html)

        return ''

    def _decode_part(self, payload: Dict[str, Any], mime_type: str) -> str:
        if payload.get('mimeType') == mime_type:
            data = payload.get('body', {}).get('data')
            if data:
                return self._decode_base64_url(data)

        for part in payload.get('parts', []):
            if part.get('mimeType') == mime_type:
                data = part.get('body', {}).get('data')
                if data:
                    return self._decode_base64_url(data)
            nested = self._decode_part(part, mime_type)
            if nested:
                return nested

        return ''

    def _decode_base64_url(self, encoded: str) -> str:
        decoded = self._decode_base64_url_bytes(encoded)
        if not decoded:
            return ''
        try:
            return decoded.decode('utf-8', errors='replace')
        except UnicodeDecodeError:
            return ''

    def _decode_base64_url_bytes(self, encoded: str) -> bytes:
        padded = encoded + '=' * ((4 - len(encoded) % 4) % 4)
        try:
            return base64.urlsafe_b64decode(padded.encode('utf-8'))
        except ValueError:
            return b''

    def _strip_html(self, value: str) -> str:
        text = re.sub(r'<[^>]+>', ' ', value)
        text = re.sub(r'\s+', ' ', text)
        return text.strip()

    def _collect_attachments(self, payload: Dict[str, Any], service: Any, message_id: str) -> List[Dict[str, Any]]:
        attachments: List[Dict[str, Any]] = []

        def walk(part: Dict[str, Any]) -> None:
            filename = str(part.get('filename', '') or '').strip()
            mime_type = str(part.get('mimeType', '') or '').strip().lower()

            if filename:
                raw_bytes = self._read_attachment_bytes(part=part, service=service, message_id=message_id)
                text_preview = self._extract_attachment_text(
                    filename=filename,
                    mime_type=mime_type,
                    raw_bytes=raw_bytes,
                )
                attachments.append(
                    {
                        'filename': filename,
                        'mime_type': mime_type,
                        'size_bytes': len(raw_bytes),
                        'text_preview': text_preview,
                    }
                )

            for nested in part.get('parts', []) or []:
                if isinstance(nested, dict):
                    walk(nested)

        walk(payload)
        return attachments

    def _read_attachment_bytes(self, part: Dict[str, Any], service: Any, message_id: str) -> bytes:
        body = part.get('body', {}) or {}
        inline_data = body.get('data')
        if isinstance(inline_data, str) and inline_data.strip():
            return self._decode_base64_url_bytes(inline_data)

        attachment_id = str(body.get('attachmentId', '') or '').strip()
        if not attachment_id:
            return b''

        try:
            attachment_payload = service.users().messages().attachments().get(
                userId='me',
                messageId=message_id,
                id=attachment_id,
            ).execute()
            encoded = str(attachment_payload.get('data', '') or '').strip()
            if not encoded:
                return b''
            return self._decode_base64_url_bytes(encoded)
        except Exception:
            return b''

    def _extract_attachment_text(self, filename: str, mime_type: str, raw_bytes: bytes) -> str:
        if not raw_bytes:
            return ''

        lower_name = filename.lower()

        if lower_name.endswith('.csv') or mime_type in {'text/csv', 'application/csv'}:
            text = self._decode_text_bytes(raw_bytes)
            if not text:
                return ''
            reader = csv.reader(StringIO(text))
            rows: List[str] = []
            for idx, row in enumerate(reader):
                rows.append(', '.join(cell.strip() for cell in row[:20]))
                if idx >= 29:
                    break
            return '\n'.join(rows).strip()

        if lower_name.endswith('.xlsx'):
            try:
                from openpyxl import load_workbook
            except ImportError:
                return ''

            try:
                workbook = load_workbook(filename=BytesIO(raw_bytes), read_only=True, data_only=True)
                sheet = workbook.active
                lines: List[str] = []
                for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                    values = [str(cell).strip() for cell in (row or []) if cell is not None]
                    if values:
                        lines.append(', '.join(values[:20]))
                    if row_idx >= 29:
                        break
                workbook.close()
                return '\n'.join(lines).strip()
            except Exception:
                return ''

        if lower_name.endswith('.pdf') or mime_type == 'application/pdf':
            try:
                from pypdf import PdfReader
            except ImportError:
                return ''

            try:
                reader = PdfReader(BytesIO(raw_bytes))
                chunks: List[str] = []
                for page in reader.pages[:3]:
                    chunks.append((page.extract_text() or '').strip())
                return '\n'.join(chunk for chunk in chunks if chunk).strip()
            except Exception:
                return ''

        if mime_type.startswith('text/') or lower_name.endswith(('.txt', '.json', '.md')):
            return self._decode_text_bytes(raw_bytes)

        return ''

    def _decode_text_bytes(self, raw_bytes: bytes) -> str:
        for encoding in ('utf-8', 'utf-16', 'latin-1'):
            try:
                return raw_bytes.decode(encoding, errors='replace')
            except Exception:
                continue
        return ''
