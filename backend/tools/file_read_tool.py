from io import BytesIO
import csv
from typing import Any, Dict, List


class FileReadTool:
    """Read and parse uploaded files into row dictionaries."""

    supported_extensions = {".csv", ".xlsx"}

    def parse_uploaded_content(self, filename: str, binary_content: bytes) -> List[Dict[str, Any]]:
        extension = '.' + filename.split('.')[-1].lower() if '.' in filename else ''
        if extension not in self.supported_extensions:
            raise ValueError('Only CSV and XLSX files are supported.')

        if extension == '.csv':
            return self._parse_csv(binary_content)
        return self._parse_xlsx(binary_content)

    def _parse_csv(self, binary_content: bytes) -> List[Dict[str, Any]]:
        text = binary_content.decode('utf-8-sig', errors='replace')
        reader = csv.DictReader(text.splitlines())
        return [dict(row) for row in reader]

    def _parse_xlsx(self, binary_content: bytes) -> List[Dict[str, Any]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError(
                'openpyxl is required for xlsx uploads. Add it to backend requirements.'
            ) from exc

        workbook = load_workbook(filename=BytesIO(binary_content), data_only=True, read_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(cell).strip() if cell is not None else '' for cell in rows[0]]
        normalized_headers = [header or f'column_{idx + 1}' for idx, header in enumerate(headers)]

        records: List[Dict[str, Any]] = []
        for row in rows[1:]:
            records.append(
                {
                    normalized_headers[idx]: row[idx] if idx < len(row) else None
                    for idx in range(len(normalized_headers))
                }
            )
        return records
